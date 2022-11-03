# import libraries
from PyPDF2 import PdfFileReader, PdfFileWriter
import re
# import xlwings as xw
import openpyxl as opxl
import pandas as pd

#looks for files based on file path
filepath = 'C:\\Users\\ccozort\\OneDrive - Bellarmine College Preparatory\\General\\Report Cards and Transcripts\\Merged\\19_20_21_22_combined.pdf'
xlFilePath = 'C:\\Users\\ccozort\\OneDrive - Bellarmine College Preparatory\\General\\Report Cards and Transcripts\\College Counseling Audits\\2022\\Class of 2023.xlsx'
# makes avaliable the files
dataframe = opxl.load_workbook(xlFilePath)
 
dataframe1 = dataframe.active
# placeholders for future lists
studentIDs = []
studentFirsts = []
studentLasts = []
# puts excel data into for loops, 1 to exclude "student id" 
for row in range(1, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, 1):
        # print(col[row].value)
        studentIDs.append(col[row].value)
for row in range(1, dataframe1.max_row):
    for col in dataframe1.iter_cols(3, 3):
        # print(col[row].value)
        studentFirsts.append(col[row].value)
for row in range(1, dataframe1.max_row):
    for col in dataframe1.iter_cols(4, 4):
        # print(col[row].value)
        studentLasts.append(col[row].value)

print(studentIDs)
print(studentFirsts)
print(studentLasts)
  
# shrtes PdfFileReader(filepath) to just pdf etc
pdf = PdfFileReader(filepath)
filewriter = PdfFileWriter()

print(pdf)
# prints # of pages
pagecount = pdf.getNumPages()

print(pagecount)


def pdfSearchExtract():
    # globalizes variables
    global filewriter
    # temporary variable
    index = 0
    # takes # of student id's and runs for that long
    for id in studentIDs:
        bookmark = 0
        match = 0
        # runs as many pages as there are-1
        for n in range(bookmark,pagecount-1):   
            print("Searching for " + str(id) + " on page " + str(n) + " of " + str(pagecount))
            # takes page
            page = pdf.getPage(n)
            # print(page)
            # stores
            Text = page.extract_text()
            # print(Text)
            # searches for id and stores as boolean
            search = re.search(str(id), Text)
            
            if search:
                bookmark = n
                match+=1
                print(search)
                filewriter = PdfFileWriter()
                filewriter.addPage(page)
                # output_filename = str(id)
                output_filename = '{}_{}_{}_{}.pdf'.format(id,studentLasts[index], studentFirsts[index], match)
                with open(output_filename,'wb') as out:
                    filewriter.write(out)
        index += 1

pdfSearchExtract()


